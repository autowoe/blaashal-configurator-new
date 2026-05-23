from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from components.models import Component, ComponentPrice, ConfigurationType

REQUIRED_COLUMNS = {"component", "parent", "order", "input_type", "group_key", "unit", "inkoop", "verkoop"}
VALID_INPUT_TYPES = {c[0] for c in Component.INPUT_TYPE_CHOICES}


def run_import(filepath: str, dry_run: bool = False) -> dict:
    """
    Importeer componenten en prijzen vanuit een Excel-bestand.
    Elke sheet = één configuratietype.

    Geeft een dict terug met:
      - total_components: int
      - total_prices: int
      - errors: list[str]
      - log: list[str]
    """
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return {"total_components": 0, "total_prices": 0, "errors": [f"Kon Excel-bestand niet openen: {e}"], "log": []}

    total_components = 0
    total_prices = 0
    errors = []
    log = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log.append(f"Verwerken sheet: '{sheet_name}'")

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            errors.append(f"[{sheet_name}] Sheet is leeg, overgeslagen.")
            continue

        headers = [str(h).strip().lower() if h else "" for h in header_row]
        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            errors.append(f"[{sheet_name}] Ontbrekende kolommen: {', '.join(sorted(missing))}")
            continue

        col_index = {name: headers.index(name) for name in REQUIRED_COLUMNS}

        rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]

        if not rows:
            log.append(f"  Geen datarijen gevonden, overgeslagen.")
            continue

        config_type = None
        if not dry_run:
            config_type, created = ConfigurationType.objects.get_or_create(name=sheet_name)
            action = "aangemaakt" if created else "gevonden"
            log.append(f"  Configuratietype {action}: '{sheet_name}'")

        sheet_components: dict[str, Component] = {}
        sheet_rows_with_errors: set[int] = set()

        for pass_num in (1, 2):
            for row_num, row in enumerate(rows, start=2):
                if row_num in sheet_rows_with_errors:
                    continue

                comp_name = _cell_str(row, col_index["component"])
                parent_name = _cell_str(row, col_index["parent"])

                if pass_num == 1 and parent_name:
                    continue
                if pass_num == 2 and not parent_name:
                    continue

                if not comp_name:
                    errors.append(f"[{sheet_name}] Rij {row_num}: 'component' is leeg, overgeslagen.")
                    sheet_rows_with_errors.add(row_num)
                    continue

                order_raw = row[col_index["order"]]
                try:
                    order = int(order_raw) if order_raw is not None else 0
                except (ValueError, TypeError):
                    errors.append(f"[{sheet_name}] Rij {row_num}: ongeldige 'order' waarde '{order_raw}'.")
                    sheet_rows_with_errors.add(row_num)
                    continue

                input_type = _cell_str(row, col_index["input_type"]) or Component.INPUT_TYPE_BOOLEAN
                if input_type not in VALID_INPUT_TYPES:
                    errors.append(
                        f"[{sheet_name}] Rij {row_num}: ongeldig input_type '{input_type}'. "
                        f"Geldige waarden: {', '.join(sorted(VALID_INPUT_TYPES))}"
                    )
                    sheet_rows_with_errors.add(row_num)
                    continue

                group_key = _cell_str(row, col_index["group_key"])
                unit = _cell_str(row, col_index["unit"])
                inkoop = _parse_decimal(row, col_index["inkoop"])
                verkoop = _parse_decimal(row, col_index["verkoop"])

                parent_obj = None
                if parent_name:
                    parent_obj = sheet_components.get(parent_name)
                    if parent_obj is None:
                        try:
                            parent_obj = Component.objects.get(name=parent_name, parent=None)
                        except Component.DoesNotExist:
                            errors.append(
                                f"[{sheet_name}] Rij {row_num}: bovenliggend component '{parent_name}' "
                                f"niet gevonden (definieer het eerst in de sheet)."
                            )
                            sheet_rows_with_errors.add(row_num)
                            continue
                        except Component.MultipleObjectsReturned:
                            errors.append(
                                f"[{sheet_name}] Rij {row_num}: meerdere componenten gevonden met naam '{parent_name}'."
                            )
                            sheet_rows_with_errors.add(row_num)
                            continue

                if not dry_run:
                    component, created = Component.objects.update_or_create(
                        name=comp_name,
                        parent=parent_obj,
                        defaults={
                            "order": order,
                            "input_type": input_type,
                            "group_key": group_key,
                            "unit": unit,
                        },
                    )
                    sheet_components[comp_name] = component
                    action = "aangemaakt" if created else "bijgewerkt"
                    log.append(f"  Component {action}: '{comp_name}'")

                    if config_type is not None:
                        ComponentPrice.objects.update_or_create(
                            configuration_type=config_type,
                            component=component,
                            defaults={"inkoop": inkoop, "verkoop": verkoop},
                        )
                        total_prices += 1
                    total_components += 1 if created else 0
                else:
                    log.append(
                        f"  [DRY-RUN] Component: '{comp_name}' "
                        f"(parent: '{parent_name or '-'}', input_type: {input_type}, "
                        f"inkoop: {inkoop}, verkoop: {verkoop})"
                    )
                    sheet_components[comp_name] = _MockComponent(comp_name, parent_name)

    return {
        "total_components": total_components,
        "total_prices": total_prices,
        "errors": errors,
        "log": log,
    }


def _cell_str(row, index) -> str:
    val = row[index]
    return str(val).strip() if val is not None else ""


def _parse_decimal(row, index) -> Decimal:
    val = row[index]
    if val is None:
        return Decimal("0.00")
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


class _MockComponent:
    def __init__(self, name, parent_name):
        self.name = name
        self.parent_name = parent_name
