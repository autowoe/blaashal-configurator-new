import json
import os
from decimal import Decimal

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = ["component", "parent", "order", "input_type", "group_key", "unit", "inkoop", "verkoop"]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

SUBCOMP_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

COLUMN_WIDTHS = {
    "component": 30,
    "parent": 25,
    "order": 8,
    "input_type": 14,
    "group_key": 18,
    "unit": 10,
    "inkoop": 14,
    "verkoop": 14,
}

INPUT_TYPE_OPTIONS = '"boolean,quantity,select,formula"'

FIXTURE_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "fixtures", "initial_data.json"
)


class Command(BaseCommand):
    help = (
        "Genereer een Excel-template voor component-import. "
        "Elke sheet = één configuratietype. "
        "Standaard pre-gevuld vanuit initial_data.json."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "output",
            nargs="?",
            default="component_template.xlsx",
            help="Uitvoerbestand (standaard: component_template.xlsx)",
        )
        parser.add_argument(
            "--empty",
            action="store_true",
            help="Genereer een leeg template zonder voorbeelddata",
        )
        parser.add_argument(
            "--from-db",
            action="store_true",
            help="Vul template vanuit de database in plaats van initial_data.json",
        )

    def handle(self, *args, **options):
        output_path = options["output"]
        use_empty = options["empty"]
        from_db = options["from_db"]

        wb = Workbook()
        wb.remove(wb.active)

        if use_empty:
            _add_empty_sheet(wb, "Configuratietype naam")
            self.stdout.write(self.style.SUCCESS(f"Leeg template opgeslagen: {output_path}"))
        elif from_db:
            _fill_from_db(wb, self.stdout)
            self.stdout.write(self.style.SUCCESS(f"Template vanuit database opgeslagen: {output_path}"))
        else:
            _fill_from_fixture(wb, self.stdout)
            self.stdout.write(self.style.SUCCESS(f"Template vanuit initial_data.json opgeslagen: {output_path}"))

        wb.save(output_path)


# ---------------------------------------------------------------------------
# Data sources
# ---------------------------------------------------------------------------

def _fill_from_fixture(wb, stdout):
    if not os.path.exists(FIXTURE_PATH):
        stdout.write(f"Fixture niet gevonden op {FIXTURE_PATH}, leeg template gegenereerd.")
        _add_empty_sheet(wb, "Voorbeeld")
        return

    with open(FIXTURE_PATH, encoding="utf-8") as f:
        data = json.load(f)

    config_types: dict[int, str] = {}
    components: dict[int, dict] = {}
    prices: list[dict] = []

    for entry in data:
        model = entry["model"]
        pk = entry["pk"]
        fields = entry["fields"]

        if model == "components.configurationtype":
            config_types[pk] = fields["name"]
        elif model == "components.component":
            components[pk] = {
                "name": fields["name"],
                "order": fields.get("order", 0),
                "input_type": fields.get("input_type", "boolean"),
                "group_key": fields.get("group_key", ""),
                "unit": fields.get("unit", ""),
                "parent": fields.get("parent"),
            }
        elif model == "components.componentprice":
            prices.append({
                "config_type": fields["configuration_type"],
                "component": fields["component"],
                "inkoop": fields.get("inkoop", "0.00"),
                "verkoop": fields.get("verkoop", "0.00"),
            })

    for ct_id, ct_name in sorted(config_types.items()):
        ct_prices = {p["component"]: p for p in prices if p["config_type"] == ct_id}
        rows = _build_rows(components, ct_prices)
        ws = wb.create_sheet(title=ct_name)
        _write_sheet(ws, rows)


def _fill_from_db(wb, stdout):
    from components.models import Component, ComponentPrice, ConfigurationType

    for config_type in ConfigurationType.objects.order_by("name"):
        components_qs = Component.objects.all()
        comp_map = {c.pk: c for c in components_qs}

        prices = {
            cp.component_id: cp
            for cp in ComponentPrice.objects.filter(configuration_type=config_type).select_related("component")
        }

        raw_components = {
            pk: {
                "name": c.name,
                "order": c.order,
                "input_type": c.input_type,
                "group_key": c.group_key,
                "unit": c.unit,
                "parent": c.parent_id,
            }
            for pk, c in comp_map.items()
        }

        raw_prices = {
            comp_id: {
                "config_type": config_type.pk,
                "component": comp_id,
                "inkoop": str(cp.inkoop or "0.00"),
                "verkoop": str(cp.verkoop or "0.00"),
            }
            for comp_id, cp in prices.items()
        }

        rows = _build_rows(raw_components, raw_prices)
        ws = wb.create_sheet(title=config_type.name)
        _write_sheet(ws, rows)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_rows(components: dict, prices: dict) -> list[tuple]:
    """Bouw gesorteerde rijen: eerst root-componenten, dan hun kinderen."""
    rows = []

    root_ids = sorted(
        [pk for pk, c in components.items() if not c["parent"]],
        key=lambda pk: components[pk]["order"],
    )

    for root_id in root_ids:
        comp = components[root_id]
        price = prices.get(root_id, {})
        rows.append((
            comp["name"],
            "",
            comp["order"],
            comp["input_type"],
            comp["group_key"],
            comp["unit"],
            price.get("inkoop", "0.00"),
            price.get("verkoop", "0.00"),
            False,  # is_subcomponent
        ))

        children = sorted(
            [pk for pk, c in components.items() if c["parent"] == root_id],
            key=lambda pk: components[pk]["order"],
        )
        for child_id in children:
            child = components[child_id]
            child_price = prices.get(child_id, {})
            rows.append((
                child["name"],
                comp["name"],
                child["order"],
                child["input_type"],
                child["group_key"],
                child["unit"],
                child_price.get("inkoop", "0.00"),
                child_price.get("verkoop", "0.00"),
                True,  # is_subcomponent
            ))

    return rows


def _add_empty_sheet(wb, sheet_name: str):
    ws = wb.create_sheet(title=sheet_name)
    _write_sheet(ws, [])


def _write_sheet(ws, rows: list[tuple]):
    # Header rij
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    # Data validatie voor input_type kolom (kolom 4)
    input_type_col = get_column_letter(HEADERS.index("input_type") + 1)
    dv = DataValidation(
        type="list",
        formula1=INPUT_TYPE_OPTIONS,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Ongeldig input type",
        error="Kies uit: boolean, quantity, select, formula",
    )
    dv.sqref = f"{input_type_col}2:{input_type_col}1048576"
    ws.add_data_validation(dv)

    # Data rijen
    for row_num, row_data in enumerate(rows, start=2):
        is_subcomp = row_data[8] if len(row_data) > 8 else False
        for col_idx, value in enumerate(row_data[:8], start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            if is_subcomp:
                cell.fill = SUBCOMP_FILL
            cell.alignment = Alignment(vertical="center")

    # Kolombreedtes
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(header, 15)

    # Bevroren koprij
    ws.freeze_panes = "A2"
